"""
Utilidades genéricas para importación masiva de terceros (clientes / proveedores)
desde archivos .xlsx o .csv, con flujo de previsualización editable y confirmación.

El mismo motor sirve para Cliente y Proveedor: cada uno declara su lista de
columnas (`COLUMNAS`) y el modelo destino, y el `TercerosImporter` se encarga de
generar la plantilla, previsualizar/validar y crear los registros.
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation

from rest_framework import status
from rest_framework.response import Response

MAX_ROWS = 500


def col(key, label, ejemplo, *, tipo='text', required=False, choices=None,
        default='', aliases=None):
    """Helper para declarar una columna de importación."""
    return {
        'key':      key,
        'label':    label,
        'ejemplo':  ejemplo,
        'tipo':     tipo,            # text | email | int | decimal | choice
        'required': required,
        'choices':  choices,         # set de valores válidos (en minúsculas) para tipo='choice'
        'default':  default,
        'aliases':  aliases or [],
    }


def _gv(fila, *keys):
    """Obtiene el primer valor no vacío entre varias claves equivalentes."""
    for k in keys:
        v = fila.get(k)
        if v is not None and str(v).strip() != '':
            return str(v).strip()
    return ''


def parsear_archivo_xlsx_csv(archivo):
    """
    Parsea un archivo .xlsx o .csv y retorna lista de dicts con claves normalizadas.
    Retorna un Response (DRF) en caso de error de formato/lectura.
    """
    nombre = archivo.name.lower()
    filas_raw = []

    try:
        if nombre.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return Response({'detail': 'El archivo está vacío.'}, status=status.HTTP_400_BAD_REQUEST)
            headers = [
                str(h).strip().lower().replace(' ', '_') if h is not None else ''
                for h in rows[0]
            ]
            for row in rows[1:]:
                fila = {h: v for h, v in zip(headers, row) if h}
                if any(v is not None and str(v).strip() != '' for v in fila.values()):
                    filas_raw.append(fila)
        elif nombre.endswith('.csv'):
            content = archivo.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            field_map = {h: h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])}
            for row in reader:
                fila = {field_map[k]: v for k, v in row.items() if k in field_map}
                if any(v and str(v).strip() for v in fila.values()):
                    filas_raw.append(fila)
        else:
            return Response(
                {'detail': 'Formato no soportado. Sube un archivo .xlsx o .csv.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as exc:  # noqa: BLE001
        return Response(
            {'detail': f'Error al leer el archivo: {exc}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not filas_raw:
        return Response({'detail': 'El archivo no contiene datos.'}, status=status.HTTP_400_BAD_REQUEST)

    return filas_raw


_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


class TercerosImporter:
    """Motor genérico de importación para un modelo de terceros."""

    def __init__(self, *, model, columnas, filename, sheet_title,
                 unique_field='nombre_comercial', defaults=None, post_init=None):
        self.model        = model
        self.columnas     = columnas
        self.filename     = filename
        self.sheet_title  = sheet_title
        self.unique_field = unique_field
        self.defaults     = defaults or {}
        # post_init(fields_dict) -> None : ajustes finales antes de crear (ej. crédito)
        self.post_init    = post_init

    # ── Plantilla .xlsx ──────────────────────────────────────────────────────
    def generar_plantilla(self):
        import openpyxl
        import openpyxl.utils
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_title

        header_fill  = PatternFill('solid', fgColor='1BAE91')
        header_font  = Font(bold=True, color='FFFFFF', size=10)
        example_fill = PatternFill('solid', fgColor='F0FDF9')
        example_font = Font(color='374151', size=10)
        border_side  = Side(style='thin', color='D1D5DB')
        thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left   = Alignment(horizontal='left', vertical='center')

        for idx, c in enumerate(self.columnas, start=1):
            hdr = ws.cell(row=1, column=idx, value=c['key'])
            hdr.font = header_font
            hdr.fill = header_fill
            hdr.alignment = center
            hdr.border = thin_border

            ex = ws.cell(row=2, column=idx, value=c['ejemplo'])
            ex.font = example_font
            ex.fill = example_fill
            ex.alignment = left
            ex.border = thin_border

            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(14, min(40, len(c['label']) + 4))

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = 'A2'

        # Hoja de instrucciones
        wi = wb.create_sheet('Instrucciones')
        wi.cell(row=1, column=1, value='Campo').font = header_font
        wi.cell(row=1, column=2, value='Obligatorio').font = header_font
        wi.cell(row=1, column=3, value='Descripción').font = header_font
        wi.cell(row=1, column=4, value='Ejemplo').font = header_font
        for cc in range(1, 5):
            cell = wi.cell(row=1, column=cc)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for r_idx, c in enumerate(self.columnas, start=2):
            desc = c['label']
            if c['tipo'] == 'choice' and c['choices']:
                desc += f". Valores: {', '.join(sorted(c['choices']))}"
            elif c['tipo'] in ('int', 'decimal'):
                desc += '. Valor numérico.'
            vals = [c['key'], 'Sí' if c['required'] else 'No', desc, str(c['ejemplo'])]
            for c_idx, val in enumerate(vals, start=1):
                cell = wi.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        wi.column_dimensions['A'].width = 22
        wi.column_dimensions['B'].width = 12
        wi.column_dimensions['C'].width = 60
        wi.column_dimensions['D'].width = 28
        wi.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        return response

    # ── Lectura de una fila cruda → valores normalizados ─────────────────────
    def _leer_fila(self, fila):
        valores = {}
        for c in self.columnas:
            raw = _gv(fila, c['key'], *c['aliases'])
            if c['tipo'] == 'choice':
                raw = raw.lower()
            valores[c['key']] = raw
        return valores

    def _validar_fila(self, valores, nombres_existentes, nombres_en_lote):
        """Retorna (estado, error_campo, mensaje)."""
        nombre = valores.get(self.unique_field, '')

        if not nombre:
            return 'error', self.unique_field, 'El nombre es obligatorio.'

        if nombre.lower() in nombres_existentes or nombre.lower() in nombres_en_lote:
            return 'omitido', None, 'Ya existe un registro con ese nombre.'

        for c in self.columnas:
            v = valores.get(c['key'], '')
            if not v:
                if c['required'] and c['key'] != self.unique_field:
                    return 'error', c['key'], f'"{c["label"]}" es obligatorio.'
                continue

            if c['tipo'] == 'choice' and c['choices'] and v.lower() not in c['choices']:
                return 'error', c['key'], f'"{v}" no válido. Opciones: {", ".join(sorted(c["choices"]))}.'
            if c['tipo'] == 'int':
                try:
                    int(float(v.replace(',', '.')))
                except (ValueError, InvalidOperation):
                    return 'error', c['key'], f'"{v}" debe ser un número entero.'
            if c['tipo'] == 'decimal':
                try:
                    Decimal(v.replace(',', '.'))
                except InvalidOperation:
                    return 'error', c['key'], f'"{v}" debe ser un número.'
            if c['tipo'] == 'email' and not _EMAIL_RE.match(v):
                return 'error', c['key'], f'"{v}" no es un correo válido.'

        return 'valido', None, None

    # ── Previsualización ─────────────────────────────────────────────────────
    def previsualizar(self, empresa, filas_raw):
        nombres_existentes = {
            n.lower() for n in self.model.objects.filter(empresa=empresa, activo=True)
            .values_list(self.unique_field, flat=True)
        }
        nombres_en_lote = set()
        resultados = []

        for i, fila in enumerate(filas_raw, start=2):
            valores = self._leer_fila(fila)
            estado, error_campo, mensaje = self._validar_fila(valores, nombres_existentes, nombres_en_lote)
            if estado == 'valido':
                nombres_en_lote.add(valores[self.unique_field].lower())
            row = dict(valores)
            row.update({'fila': i, 'estado': estado, 'error_campo': error_campo, 'mensaje': mensaje})
            resultados.append(row)

        return {
            'total':       len(resultados),
            'validos':     sum(1 for r in resultados if r['estado'] == 'valido'),
            'con_errores': sum(1 for r in resultados if r['estado'] == 'error'),
            'omitidos':    sum(1 for r in resultados if r['estado'] == 'omitido'),
            'filas':       resultados,
        }

    # ── Construir kwargs del modelo a partir de una fila ─────────────────────
    def _fields_para_crear(self, valores):
        fields = {}
        for c in self.columnas:
            v = valores.get(c['key'], '')
            if c['tipo'] == 'int':
                fields[c['key']] = int(float(v.replace(',', '.'))) if v else 0
            elif c['tipo'] == 'decimal':
                fields[c['key']] = Decimal(v.replace(',', '.')) if v else Decimal('0')
            elif c['tipo'] == 'choice':
                fields[c['key']] = v.lower() if v else (c['default'] or '')
            else:
                fields[c['key']] = v if v else (c['default'] or '')
        for k, default in self.defaults.items():
            if not fields.get(k):
                fields[k] = default
        if self.post_init:
            self.post_init(fields)
        return fields

    # ── Importación ──────────────────────────────────────────────────────────
    def importar(self, empresa, filas_raw, modo):
        from django.db import transaction

        nombres_existentes = {
            n.lower() for n in self.model.objects.filter(empresa=empresa, activo=True)
            .values_list(self.unique_field, flat=True)
        }
        nombres_en_lote = set()
        resultados = []
        a_crear = []

        for i, fila in enumerate(filas_raw, start=2):
            fila_num = int(fila['fila']) if str(fila.get('fila', '')).strip().isdigit() else i
            valores = self._leer_fila(fila)
            estado, _campo, mensaje = self._validar_fila(valores, nombres_existentes, nombres_en_lote)
            nombre = valores.get(self.unique_field, '') or '(vacío)'

            if estado == 'valido':
                nombres_en_lote.add(valores[self.unique_field].lower())
                a_crear.append((fila_num, nombre, valores))
            else:
                resultados.append({'fila': fila_num, 'nombre': nombre, 'estado': estado, 'mensaje': mensaje})

        # Modo atómico: abortar si hay errores de validación
        if modo == 'atomico':
            errores_val = [r for r in resultados if r['estado'] == 'error']
            if errores_val:
                return {
                    'creados':  0,
                    'omitidos': sum(1 for r in resultados if r['estado'] == 'omitido'),
                    'errores':  len(errores_val),
                    'filas':    sorted(resultados, key=lambda x: x['fila']),
                }, status.HTTP_422_UNPROCESSABLE_ENTITY

        creados = 0

        def _crear_uno(item):
            nonlocal creados
            fila_num, nombre, valores = item
            fields = self._fields_para_crear(valores)
            self.model.objects.create(empresa=empresa, **fields)
            resultados.append({'fila': fila_num, 'nombre': nombre, 'estado': 'creado'})
            creados += 1

        if modo == 'atomico':
            with transaction.atomic():
                for item in a_crear:
                    _crear_uno(item)
        else:
            for item in a_crear:
                try:
                    with transaction.atomic():
                        _crear_uno(item)
                except Exception as exc:  # noqa: BLE001
                    resultados.append({'fila': item[0], 'nombre': item[1], 'estado': 'error', 'mensaje': str(exc)})

        return {
            'creados':  creados,
            'omitidos': sum(1 for r in resultados if r['estado'] == 'omitido'),
            'errores':  sum(1 for r in resultados if r['estado'] == 'error'),
            'filas':    sorted(resultados, key=lambda x: x['fila']),
        }, status.HTTP_200_OK
